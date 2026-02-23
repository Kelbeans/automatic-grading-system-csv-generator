"""
SF10 Grade Automation System
Generates individual SF10 files for each student with their 1st quarter grades
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
from pathlib import Path


class SF10Generator:
    """Handles the generation of SF10 files for students"""

    def __init__(self, grading_sheet_path, sf10_template_path, output_dir='output', learners_profile_path=None):
        """
        Initialize the SF10 Generator

        Args:
            grading_sheet_path: Path to the grading sheet Excel file
            sf10_template_path: Path to the SF10 template Excel file
            output_dir: Directory where generated SF10 files will be saved
            learners_profile_path: Optional path to learners profile file with LRN, Birthday, Sex
        """
        self.grading_sheet_path = grading_sheet_path
        self.sf10_template_path = sf10_template_path
        self.output_dir = output_dir
        self.learners_profile_path = learners_profile_path

        # Only load learners profile if explicitly provided
        self.learners_profile_data = self._load_learners_profile() if learners_profile_path else {}

        # Create output directory if it doesn't exist
        Path(self.output_dir).mkdir(exist_ok=True)

        # Define subject mapping between SUMMARY sheet and SF10
        self.subject_mapping = {
            'LANGUAGE': 'Language',
            'READING AND LITERACY': 'Reading and Literacy',
            'MATH': 'Mathematics',
            'GMRC': 'GMRC (Good Manners and Right Conduct)',
            'MAKABANSA': 'Makabansa'
        }

        # SF10 row positions for each subject (0-indexed)
        self.sf10_subject_rows = {
            'Language': 29,
            'Reading and Literacy': 30,
            'Mathematics': 31,
            'GMRC (Good Manners and Right Conduct)': 32,
            'Makabansa': 33
        }

        # Column positions in SF10 (0-indexed)
        self.sf10_quarter_columns = {
            1: 10,  # 1st quarter in column K (index 10)
            2: 11,  # 2nd quarter in column L (index 11)
            3: 12,  # 3rd quarter in column M (index 12)
            4: 13   # 4th quarter in column N (index 13)
        }

    def _normalize_name(self, name):
        """
        Normalize student name for matching
        - Convert to uppercase
        - Remove extra spaces
        - Normalize comma spacing (remove spaces around commas, then add one space after each)
        """
        # Remove extra whitespace
        name = ' '.join(name.split())
        # Remove spaces around commas
        name = name.replace(' ,', ',').replace(', ', ',')
        # Add consistent space after each comma
        name = name.replace(',', ', ')
        return name.strip().upper()

    def _load_learners_profile(self):
        """
        Load learner profile data (LRN, Birthday, Sex) from LEARNERS PROFILE.xlsx

        Returns:
            Dictionary mapping student names to their profile data
        """
        if not os.path.exists(self.learners_profile_path):
            print(f"Warning: Learners profile not found at {self.learners_profile_path}")
            return {}

        try:
            # Read the profile file (header at row 3, data starts at row 4)
            df = pd.read_excel(self.learners_profile_path, header=None, skiprows=4)

            # Extract relevant columns: 1=LRN, 4=NAME, 8=BIRTHDAY, 10=SEX
            df_clean = df[[1, 4, 8, 10]]
            df_clean.columns = ['LRN', 'NAME', 'BIRTHDAY', 'SEX']

            # Remove any NaN rows
            df_clean = df_clean.dropna(subset=['NAME'])

            # Create dictionary mapping name to profile data
            profile_dict = {}
            for _, row in df_clean.iterrows():
                name = self._normalize_name(str(row['NAME']))
                profile_dict[name] = {
                    'LRN': row['LRN'],
                    'BIRTHDAY': row['BIRTHDAY'],
                    'SEX': str(row['SEX']).strip().upper()
                }

            print(f"✓ Loaded {len(profile_dict)} learner profiles")
            return profile_dict

        except Exception as e:
            print(f"Warning: Could not load learners profile: {e}")
            return {}

    def read_student_grades(self):
        """
        Read student data and grades from the SUMMARY sheet

        Returns:
            List of dictionaries containing student information and grades
        """
        # Read the SUMMARY sheet
        df = pd.read_excel(
            self.grading_sheet_path,
            sheet_name='SUMMARY OF QUARTERLY GRADES ',
            header=None
        )

        students = []

        # Start reading from row 10 (0-indexed: row 10 is the first student)
        # Row 7 has headers, row 9 has "MALE", row 10 starts student data
        start_row = 10

        for idx in range(start_row, len(df)):
            row = df.iloc[idx]

            # Check if we have valid student data
            if pd.isna(row[1]) or row[1] == '':
                continue

            # Extract student name (column 1)
            student_name = str(row[1]).strip()

            # Skip if this is a header row or empty
            if student_name in ['MALE', 'FEMALE', "LEARNERS' NAMES", '']:
                continue

            # Extract grades from columns 5-9 with safe access
            def safe_get_grade(row, col_idx):
                """Safely get grade value from row, handling missing columns"""
                try:
                    value = row[col_idx] if col_idx in row.index else 0
                    return value if not pd.isna(value) else 0
                except (KeyError, IndexError):
                    return 0

            grades = {
                'LANGUAGE': safe_get_grade(row, 5),
                'READING AND LITERACY': safe_get_grade(row, 6),
                'MATH': safe_get_grade(row, 7),
                'GMRC': safe_get_grade(row, 8),
                'MAKABANSA': safe_get_grade(row, 9)
            }

            students.append({
                'name': student_name,
                'grades': grades
            })

        return students

    def generate_sf10_for_student(self, student, quarter=1):
        """
        Generate an SF10 file for a single student

        Args:
            student: Dictionary containing student name and grades
            quarter: Quarter number (1-4, default is 1)
        """
        # Load the SF10 template
        wb = load_workbook(self.sf10_template_path)
        ws = wb.active

        # Parse student name (format: "LAST,FIRST, MIDDLE")
        student_name = student['name']
        name_parts = student_name.split(',')
        last_name = name_parts[0].strip() if len(name_parts) > 0 else ''
        first_name = name_parts[1].strip() if len(name_parts) > 1 else ''
        middle_name = name_parts[2].strip() if len(name_parts) > 2 else ''

        # Update the name fields in the SF10
        # In openpyxl, rows and columns are 1-indexed
        # Row 9, Column E (5): Last Name
        ws.cell(row=9, column=5, value=last_name)

        # Row 9, Column Q (17): First Name
        ws.cell(row=9, column=17, value=first_name)

        # Row 9, Column AP (42): Middle Name
        ws.cell(row=9, column=42, value=middle_name)

        # Fill in LRN, Birthday, Sex from learners profile if available
        student_name_upper = self._normalize_name(student_name)
        if student_name_upper in self.learners_profile_data:
            profile = self.learners_profile_data[student_name_upper]

            # Row 10, Column J (10): LRN - format as text to prevent scientific notation
            lrn_cell = ws.cell(row=10, column=10)
            lrn_cell.value = str(int(profile['LRN'])) if pd.notna(profile['LRN']) else ''
            lrn_cell.number_format = '@'  # Text format

            # Row 10, Column U (21): Birthday
            ws.cell(row=10, column=21, value=profile['BIRTHDAY'])

            # Row 10, Column AS (45): Sex
            ws.cell(row=10, column=45, value=profile['SEX'])

        # Fill in the grades for each subject in the appropriate quarter column
        quarter_col = self.sf10_quarter_columns[quarter]

        for summary_subject, sf10_subject in self.subject_mapping.items():
            grade = student['grades'].get(summary_subject, 0)
            row_idx = self.sf10_subject_rows[sf10_subject]

            # openpyxl uses 1-indexed rows/columns
            ws.cell(row=row_idx + 1, column=quarter_col + 1, value=grade)

            # Clear other quarter columns (only fill the specified quarter)
            for q in range(1, 5):
                if q != quarter:
                    other_col = self.sf10_quarter_columns[q]
                    # Use empty string to clear the cell (None doesn't work in openpyxl)
                    ws.cell(row=row_idx + 1, column=other_col + 1).value = ''

        # Generate output filename
        # Clean the name for filename (remove commas, spaces, special chars)
        safe_name = last_name.replace(',', '').replace(' ', '_')
        output_filename = f'SF10_{safe_name}_{first_name.replace(" ", "_")}.xlsx'
        output_path = os.path.join(self.output_dir, output_filename)

        # Save the workbook
        wb.save(output_path)

        return output_path

    def generate_all_sf10s(self, quarter=1):
        """
        Generate SF10 files for all students

        Args:
            quarter: Quarter number (1-4, default is 1)

        Returns:
            List of generated file paths
        """
        print(f'Reading student data from: {self.grading_sheet_path}')
        students = self.read_student_grades()

        print(f'\nFound {len(students)} students')
        print(f'Generating SF10 files for Quarter {quarter}...\n')

        generated_files = []

        for idx, student in enumerate(students, 1):
            try:
                output_path = self.generate_sf10_for_student(student, quarter)
                print(f'{idx}. Generated: {os.path.basename(output_path)} - {student["name"]}')
                generated_files.append(output_path)
            except Exception as e:
                print(f'{idx}. ERROR generating SF10 for {student["name"]}: {str(e)}')

        print(f'\n✅ Successfully generated {len(generated_files)} SF10 files in "{self.output_dir}" directory')

        return generated_files

    def generate_single_workbook_all_students(self, quarter=1, output_filename='SF10_All_Students.xlsx'):
        """
        Generate a single Excel workbook with one sheet per student
        Preserves images/logos from template by copying the entire workbook structure

        Args:
            quarter: Quarter number (1-4, default is 1)
            output_filename: Name of the output file

        Returns:
            Path to the generated workbook
        """
        import shutil
        from copy import copy

        print(f'Reading student data from: {self.grading_sheet_path}')
        students = self.read_student_grades()

        print(f'\nFound {len(students)} students')
        print(f'Generating single workbook with {len(students)} tabs for Quarter {quarter}...\n')

        # Create output path
        output_path = os.path.join(self.output_dir, output_filename)

        # Copy the template file to preserve all media/drawings
        shutil.copy(self.sf10_template_path, output_path)

        # Load the copied workbook
        output_wb = load_workbook(output_path)
        template_ws = output_wb.active  # This is the template sheet with images

        for idx, student in enumerate(students, 1):
            try:
                # Parse student name for sheet name
                student_name = student['name']
                name_parts = student_name.split(',')
                last_name = name_parts[0].strip() if len(name_parts) > 0 else ''
                first_name = name_parts[1].strip() if len(name_parts) > 1 else ''
                middle_name = name_parts[2].strip() if len(name_parts) > 2 else ''

                # Create a clean sheet name (max 31 chars, no special chars)
                sheet_name = f"{last_name[:15]} {first_name[:10]}".strip()
                # Remove invalid characters for sheet names
                for char in [':', '\\', '/', '?', '*', '[', ']']:
                    sheet_name = sheet_name.replace(char, '')

                # Copy the template sheet
                new_ws = output_wb.copy_worksheet(template_ws)
                new_ws.title = sheet_name

                # Add logos to the new sheet
                try:
                    from openpyxl.drawing.image import Image as XLImage
                    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                    from openpyxl.drawing.xdr import XDRPositiveSize2D
                    from openpyxl.utils.units import pixels_to_EMU

                    # Path to logo files (PNG with transparent backgrounds)
                    project_root = os.path.dirname(os.path.abspath(__file__))
                    kagawaran_seal_path = os.path.join(project_root, 'assets', 'img', 'c128a1b0-b3b3-492b-84f3-6624923eb8b4-removebg-preview.png')
                    deped_logo_path = os.path.join(project_root, 'assets', 'img', '95c51f57-dfdc-418a-bf4c-54acb45308be-removebg-preview.png')

                    # Add Kagawaran ng Edukasyon seal (left)
                    # Excel properties: Width=87pt, Height=90pt, Position X=43pt, Y=0pt
                    if os.path.exists(kagawaran_seal_path):
                        kagawaran_seal = XLImage(kagawaran_seal_path)
                        # Size in pixels (1pt = 1.33 pixels)
                        width_px = int(87 * 1.33)   # 116 pixels
                        height_px = int(90 * 1.33)  # 120 pixels

                        # Position in pixels
                        x_px = int(43 * 1.33)  # 57 pixels from left
                        y_px = 0  # 0 pixels from top

                        # Create anchor with position and size
                        marker = AnchorMarker(col=0, colOff=pixels_to_EMU(x_px),
                                             row=0, rowOff=pixels_to_EMU(y_px))
                        size = XDRPositiveSize2D(pixels_to_EMU(width_px), pixels_to_EMU(height_px))
                        kagawaran_seal.anchor = OneCellAnchor(_from=marker, ext=size)
                        new_ws.add_image(kagawaran_seal)

                    # Add DepED logo (right)
                    # Excel properties: Width=137pt, Height=137pt, Position X=821pt, Y=-24pt
                    if os.path.exists(deped_logo_path):
                        deped_logo = XLImage(deped_logo_path)
                        # Size in pixels (square, constrain proportions)
                        width_px = int(137 * 1.33)   # 182 pixels
                        height_px = int(137 * 1.33)  # 182 pixels

                        # Position in points (1 point = 914400 EMU)
                        # Using exact point values for accurate positioning
                        x_pt = 821
                        y_pt = -24

                        # Convert points directly to EMU (1 pt = 12700 EMU)
                        x_emu = int(x_pt * 12700)
                        y_emu = int(y_pt * 12700)

                        # Create anchor with position and size
                        marker = AnchorMarker(col=0, colOff=x_emu,
                                             row=0, rowOff=y_emu)
                        size = XDRPositiveSize2D(pixels_to_EMU(width_px), pixels_to_EMU(height_px))
                        deped_logo.anchor = OneCellAnchor(_from=marker, ext=size)
                        new_ws.add_image(deped_logo)
                except Exception as e:
                    print(f'   Warning: Could not add logos to {sheet_name}: {str(e)}')
                    # Continue without logos if there's an issue

                # Now fill in the student data
                # Update the name fields
                new_ws.cell(row=9, column=5, value=last_name)
                new_ws.cell(row=9, column=17, value=first_name)
                new_ws.cell(row=9, column=42, value=middle_name)

                # Fill in LRN, Birthday, Sex from learners profile if available
                student_name_upper = self._normalize_name(student_name)
                print(f"   DEBUG: Looking for '{student_name_upper}' in profile data (has {len(self.learners_profile_data)} entries)")
                if student_name_upper in self.learners_profile_data:
                    profile = self.learners_profile_data[student_name_upper]
                    print(f"   ✓ Found profile: LRN={profile['LRN']}, Birthday={profile['BIRTHDAY']}, Sex={profile['SEX']}")

                    # Row 10, Column J (10): LRN - format as text to prevent scientific notation
                    lrn_cell = new_ws.cell(row=10, column=10)
                    lrn_cell.value = str(int(profile['LRN'])) if pd.notna(profile['LRN']) else ''
                    lrn_cell.number_format = '@'  # Text format

                    # Row 10, Column U (21): Birthday
                    new_ws.cell(row=10, column=21, value=profile['BIRTHDAY'])

                    # Row 10, Column AS (45): Sex
                    new_ws.cell(row=10, column=45, value=profile['SEX'])
                else:
                    print(f"   ✗ NOT FOUND in profile data")

                # Fill in the grades for the specified quarter
                quarter_col = self.sf10_quarter_columns[quarter]

                for summary_subject, sf10_subject in self.subject_mapping.items():
                    grade = student['grades'].get(summary_subject, 0)
                    row_idx = self.sf10_subject_rows[sf10_subject]

                    # Fill the grade
                    new_ws.cell(row=row_idx + 1, column=quarter_col + 1, value=grade)

                    # Clear other quarter columns
                    for q in range(1, 5):
                        if q != quarter:
                            other_col = self.sf10_quarter_columns[q]
                            new_ws.cell(row=row_idx + 1, column=other_col + 1).value = ''

                print(f'{idx}. Added sheet: {sheet_name} - {student["name"]}')

            except Exception as e:
                print(f'{idx}. ERROR adding sheet for {student["name"]}: {str(e)}')

        # Remove the original template sheet (it was only used for copying)
        output_wb.remove(template_ws)

        # Save the workbook
        output_wb.save(output_path)
        print('   ✓ SF10 generated with logos')

        print(f'\n✅ Successfully generated single workbook with {len(output_wb.sheetnames)} sheets')
        print(f'   Saved to: {output_path}')

        return output_path


def main():
    """
    Main function to run the SF10 generation

    Usage:
        python generate_sf10.py your_grading_sheet.xlsx

    Or modify this function to point to your grading sheet file.
    """
    import sys

    if len(sys.argv) < 2:
        print("Usage: python generate_sf10.py <grading_sheet.xlsx>")
        print("\nExample:")
        print("  python generate_sf10.py '1st Quarter Grades.xlsx'")
        print("\nOr use the web interface:")
        print("  python sf10_web_app.py")
        sys.exit(1)

    grading_sheet = sys.argv[1]
    sf10_template = 'assets/docs/SF10.xlsx'

    if not os.path.exists(grading_sheet):
        print(f"Error: File not found: {grading_sheet}")
        sys.exit(1)

    # Create generator instance
    generator = SF10Generator(
        grading_sheet_path=grading_sheet,
        sf10_template_path=sf10_template,
        output_dir='output'
    )

    # Generate single workbook with all students for 1st quarter
    generator.generate_single_workbook_all_students(quarter=1, output_filename='SF10_All_Students_Q1.xlsx')


if __name__ == '__main__':
    main()
