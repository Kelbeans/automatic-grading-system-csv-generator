"""
SF10 Grade Automation - Web Interface
Flask web application with drag-and-drop for non-technical users
"""

from flask import Flask, render_template, request, send_file, jsonify
from flask_cors import CORS
import os
from werkzeug.utils import secure_filename
from generate_sf10 import SF10Generator
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
import tempfile
import shutil

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes
app.config['UPLOAD_FOLDER'] = tempfile.mkdtemp()
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Template path
TEMPLATE_PATH = 'assets/docs/SF10.xlsx'


def identify_quarter_from_filename(filename):
    """Identify which quarter from filename"""
    filename_lower = filename.lower()
    if '1st' in filename_lower or '1 ' in filename_lower or 'first' in filename_lower:
        return 1
    elif '2nd' in filename_lower or '2 ' in filename_lower or 'second' in filename_lower:
        return 2
    elif '3rd' in filename_lower or '3 ' in filename_lower or 'third' in filename_lower:
        return 3
    elif '4th' in filename_lower or '4 ' in filename_lower or 'fourth' in filename_lower:
        return 4
    return None


def is_sf10_file(filepath):
    """
    Check if a file is an SF10 by examining its structure

    Returns True if:
    - File has MANY sheets (20+, one per student)
    - Sheet names are short (student names)
    - Has expected SF10 structure (row 9 has name fields, rows 30-34 have grades)
    - Has many merged cells (SF10 has ~500 merged cells)
    """
    try:
        wb = load_workbook(filepath)

        # Must have MANY sheets (one per student) - at least 20
        # Grading sheets typically have 5-7 sheets (subjects + summary)
        if len(wb.sheetnames) < 20:
            return False

        # Check first sheet for SF10 structure
        first_sheet = wb[wb.sheetnames[0]]

        # Check if row 9, column E (5) specifically has a student last name
        # This is more specific than checking multiple columns
        last_name_cell = first_sheet.cell(row=9, column=5).value
        if not (last_name_cell and isinstance(last_name_cell, str) and len(last_name_cell.strip()) > 0):
            return False

        # Check if rows 30-34 column B have subject names (Language, Reading, Math)
        # This is specific to SF10 grade section
        subject_indicators = []
        for row in [30, 31, 32]:
            val = first_sheet.cell(row=row, column=2).value
            if val and isinstance(val, str):
                subject_indicators.append(val.lower())

        # Should have words like "language", "reading", "math" in these cells
        has_subjects = any(
            word in ' '.join(subject_indicators)
            for word in ['language', 'reading', 'math']
        )

        # Check for merged cells (SF10 has many, grading sheets don't)
        has_many_merged = len(first_sheet.merged_cells.ranges) > 100

        # All conditions must be True
        return bool(last_name_cell and has_subjects and has_many_merged)

    except Exception as e:
        print(f"Error checking if file is SF10: {e}")
        return False


def merge_quarters_into_sf10(grading_files, existing_sf10_path=None, output_path=None):
    """
    Merge grades from multiple quarters into SF10

    Args:
        grading_files: List of tuples (filepath, quarter_number)
        existing_sf10_path: Path to existing SF10 to update (optional)
        output_path: Where to save the result

    Returns:
        Path to generated file
    """
    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f'SF10_All_Students_{timestamp}.xlsx')

    # If we have an existing SF10, use it as base
    if existing_sf10_path and os.path.exists(existing_sf10_path):
        print(f"Using existing SF10 as base: {existing_sf10_path}")
        # Copy it to output path as starting point
        shutil.copy(existing_sf10_path, output_path)
        base_wb = load_workbook(output_path)
    else:
        print("Creating fresh SF10 from template")
        # Create fresh from first grading file
        first_file, first_quarter = grading_files[0]
        generator = SF10Generator(
            grading_sheet_path=first_file,
            sf10_template_path=TEMPLATE_PATH,
            output_dir=app.config['UPLOAD_FOLDER']
        )
        temp_output = generator.generate_single_workbook_all_students(
            quarter=first_quarter,
            output_filename=os.path.basename(output_path)
        )
        base_wb = load_workbook(temp_output)

        # Remove the first file since we already processed it
        grading_files = grading_files[1:]

    # Now update with remaining quarters
    for grading_file, quarter in grading_files:
        print(f"Adding Quarter {quarter} from {grading_file}")

        # Read students from this grading sheet
        generator = SF10Generator(
            grading_sheet_path=grading_file,
            sf10_template_path=TEMPLATE_PATH,
            output_dir=app.config['UPLOAD_FOLDER']
        )
        students = generator.read_student_grades()

        # Update each student's sheet with this quarter's grades
        for student in students:
            # Find the matching sheet (by last name)
            name_parts = student['name'].split(',')
            last_name = name_parts[0].strip()
            first_name = name_parts[1].strip() if len(name_parts) > 1 else ''

            # Try to find matching sheet
            sheet_name_pattern = f"{last_name[:15]} {first_name[:10]}".strip()
            matching_sheet = None

            for sheet_name in base_wb.sheetnames:
                if last_name in sheet_name:
                    matching_sheet = sheet_name
                    break

            if matching_sheet:
                ws = base_wb[matching_sheet]

                # Fill in grades for this quarter
                quarter_col = generator.sf10_quarter_columns[quarter]

                for summary_subject, sf10_subject in generator.subject_mapping.items():
                    grade = student['grades'].get(summary_subject, 0)
                    row_idx = generator.sf10_subject_rows[sf10_subject]
                    ws.cell(row=row_idx + 1, column=quarter_col + 1, value=grade)

                print(f"  Updated: {matching_sheet}")
            else:
                print(f"  Warning: Could not find sheet for {student['name']}")

    # Save the updated workbook
    base_wb.save(output_path)
    print(f"Saved to: {output_path}")

    return output_path


@app.route('/')
def index():
    """Main page with drag-and-drop interface"""
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle file uploads and process"""
    try:
        # Get uploaded files
        grading_files = []
        existing_sf10 = None

        # Check for grading sheets
        for key in request.files:
            file = request.files[key]
            if file and file.filename:
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                # Check if this is an existing SF10 or a grading sheet
                # First check by file structure (more reliable)
                if is_sf10_file(filepath):
                    existing_sf10 = filepath
                    print(f"Detected existing SF10 (by structure): {filename}")
                else:
                    # Not an SF10, so it must be a grading sheet
                    # Try to identify quarter from filename
                    quarter = identify_quarter_from_filename(filename)
                    if quarter:
                        grading_files.append((filepath, quarter))
                        print(f"Detected grading sheet: {filename} -> Quarter {quarter}")
                    else:
                        # Check if this might be an SF10 with unusual naming
                        return jsonify({
                            'error': f'Could not identify file type: {filename}. '
                                   f'Grading sheets should include "1st", "2nd", "3rd", or "4th" in the filename. '
                                   f'SF10 files should have multiple student sheets with grades.'
                        }), 400

        if not grading_files:
            return jsonify({'error': 'No grading sheets uploaded'}), 400

        # Sort by quarter
        grading_files.sort(key=lambda x: x[1])

        # Generate/update SF10
        output_path = merge_quarters_into_sf10(grading_files, existing_sf10)

        # Prepare response
        quarters_processed = [q for _, q in grading_files]

        return jsonify({
            'success': True,
            'message': f'Successfully processed quarters: {quarters_processed}',
            'download_url': f'/download/{os.path.basename(output_path)}',
            'filename': os.path.basename(output_path),
            'quarters': quarters_processed
        })

    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/download/<filename>')
def download_file(filename):
    """Download the generated SF10 file"""
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    if os.path.exists(filepath):
        return send_file(filepath, as_attachment=True, download_name='SF10_All_Students.xlsx')
    return "File not found", 404


if __name__ == '__main__':
    print("\n" + "="*70)
    print("SF10 Grade Automation Web Interface")
    print("="*70)
    print("\nStarting server...")
    print("Open your browser to: http://localhost:8080")
    print("\nPress CTRL+C to stop the server")
    print("="*70 + "\n")

    app.run(debug=True, port=8080, host='127.0.0.1')
