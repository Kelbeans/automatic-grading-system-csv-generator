"""
Unit tests for SF10 Grade Automation System
"""

import unittest
import os
import tempfile
import shutil
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
import pandas as pd
from openpyxl import Workbook, load_workbook

from generate_sf10 import SF10Generator


class TestSF10Generator(unittest.TestCase):
    """Test cases for SF10Generator class"""

    def setUp(self):
        """Set up test fixtures"""
        # Create temporary directory for test files
        self.test_dir = tempfile.mkdtemp()

        # Create mock grading sheet
        self.grading_sheet_path = os.path.join(self.test_dir, 'test_grading.xlsx')
        self.sf10_template_path = os.path.join(self.test_dir, 'test_sf10.xlsx')
        self.output_dir = os.path.join(self.test_dir, 'output')

        # Create a simple grading sheet for testing
        self._create_mock_grading_sheet()

        # Create a simple SF10 template for testing
        self._create_mock_sf10_template()

    def tearDown(self):
        """Clean up test fixtures"""
        # Remove temporary directory
        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)

    def _create_mock_grading_sheet(self):
        """Create a mock grading sheet Excel file"""
        wb = Workbook()
        ws = wb.create_sheet('SUMMARY OF QUARTERLY GRADES ')

        # Add header rows (simplified version)
        ws.cell(row=8, column=2, value="LEARNERS' NAMES")
        ws.cell(row=8, column=6, value='LANGUAGE')
        ws.cell(row=8, column=7, value='READING AND LITERACY')
        ws.cell(row=8, column=8, value='MATH')
        ws.cell(row=8, column=9, value='GMRC')
        ws.cell(row=8, column=10, value='MAKABANSA')

        # Add student data starting from row 11 (Excel uses 1-indexing)
        students = [
            ('AGOT,KHIAN CLOUD, DABLO', 83, 81, 86, 85, 85),
            ('ANDEO,JHON PAUL, ANITADO', 84, 82, 85, 86, 86),
            ('ANTONIO,ZAMUEL ELLISE, NAVARRO', 83, 85, 83, 84, 85)
        ]

        for idx, (name, lang, read, math, gmrc, maka) in enumerate(students, start=11):
            ws.cell(row=idx, column=2, value=name)
            ws.cell(row=idx, column=6, value=lang)
            ws.cell(row=idx, column=7, value=read)
            ws.cell(row=idx, column=8, value=math)
            ws.cell(row=idx, column=9, value=gmrc)
            ws.cell(row=idx, column=10, value=maka)

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        wb.save(self.grading_sheet_path)

    def _create_mock_sf10_template(self):
        """Create a mock SF10 template Excel file"""
        wb = Workbook()
        ws = wb.active

        # Add basic structure
        ws.cell(row=9, column=5, value='LAST NAME')
        ws.cell(row=28, column=11, value='Quarterly Rating')
        ws.cell(row=29, column=11, value=1)
        ws.cell(row=29, column=12, value=2)
        ws.cell(row=29, column=13, value=3)
        ws.cell(row=29, column=14, value=4)

        # Add subject names
        ws.cell(row=30, column=2, value='Language')
        ws.cell(row=31, column=2, value='Reading and Literacy')
        ws.cell(row=32, column=2, value='Mathematics')
        ws.cell(row=33, column=2, value='GMRC (Good Manners and Right Conduct)')
        ws.cell(row=34, column=2, value='Makabansa')

        wb.save(self.sf10_template_path)

    def test_initialization(self):
        """Test SF10Generator initialization"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        self.assertEqual(generator.grading_sheet_path, self.grading_sheet_path)
        self.assertEqual(generator.sf10_template_path, self.sf10_template_path)
        self.assertEqual(generator.output_dir, self.output_dir)
        self.assertTrue(os.path.exists(self.output_dir))

    def test_subject_mapping_defined(self):
        """Test that subject mapping is properly defined"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        expected_subjects = ['LANGUAGE', 'READING AND LITERACY', 'MATH', 'GMRC', 'MAKABANSA']
        for subject in expected_subjects:
            self.assertIn(subject, generator.subject_mapping)

    def test_sf10_subject_rows_defined(self):
        """Test that SF10 subject row positions are defined"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        expected_sf10_subjects = [
            'Language',
            'Reading and Literacy',
            'Mathematics',
            'GMRC (Good Manners and Right Conduct)',
            'Makabansa'
        ]

        for subject in expected_sf10_subjects:
            self.assertIn(subject, generator.sf10_subject_rows)
            self.assertIsInstance(generator.sf10_subject_rows[subject], int)

    def test_sf10_quarter_columns_defined(self):
        """Test that quarter column positions are defined"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        for quarter in range(1, 5):
            self.assertIn(quarter, generator.sf10_quarter_columns)
            self.assertIsInstance(generator.sf10_quarter_columns[quarter], int)

    def test_read_student_grades(self):
        """Test reading student grades from grading sheet"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        students = generator.read_student_grades()

        # Should have 3 students
        self.assertEqual(len(students), 3)

        # Check first student
        first_student = students[0]
        self.assertEqual(first_student['name'], 'AGOT,KHIAN CLOUD, DABLO')
        self.assertEqual(first_student['grades']['LANGUAGE'], 83)
        self.assertEqual(first_student['grades']['READING AND LITERACY'], 81)
        self.assertEqual(first_student['grades']['MATH'], 86)
        self.assertEqual(first_student['grades']['GMRC'], 85)
        self.assertEqual(first_student['grades']['MAKABANSA'], 85)

    def test_read_student_grades_with_missing_data(self):
        """Test reading student grades handles missing data"""
        # Create a grading sheet with missing grades
        wb = Workbook()
        ws = wb.create_sheet('SUMMARY OF QUARTERLY GRADES ')

        # Add student with missing grades
        ws.cell(row=11, column=2, value='TEST,STUDENT, A')
        ws.cell(row=11, column=6, value=80)
        # Leave other grades empty

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        temp_grading = os.path.join(self.test_dir, 'test_missing.xlsx')
        wb.save(temp_grading)

        generator = SF10Generator(
            temp_grading,
            self.sf10_template_path,
            self.output_dir
        )

        students = generator.read_student_grades()
        self.assertEqual(len(students), 1)

        # Missing grades should default to 0
        student = students[0]
        self.assertEqual(student['grades']['LANGUAGE'], 80)
        self.assertEqual(student['grades']['READING AND LITERACY'], 0)
        self.assertEqual(student['grades']['MATH'], 0)

    def test_generate_sf10_for_student(self):
        """Test generating SF10 for a single student"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        student = {
            'name': 'DOE,JOHN, MIDDLE',
            'grades': {
                'LANGUAGE': 85,
                'READING AND LITERACY': 87,
                'MATH': 90,
                'GMRC': 88,
                'MAKABANSA': 86
            }
        }

        output_path = generator.generate_sf10_for_student(student, quarter=1)

        # Check file was created
        self.assertTrue(os.path.exists(output_path))

        # Verify content
        wb = load_workbook(output_path)
        ws = wb.active

        # Check last name was filled (row 9, column 5)
        self.assertEqual(ws.cell(row=9, column=5).value, 'DOE')

        # Check first name was filled (row 9, column 17)
        self.assertEqual(ws.cell(row=9, column=17).value, 'JOHN')

        # Check middle name was filled (row 9, column 42)
        self.assertEqual(ws.cell(row=9, column=42).value, 'MIDDLE')

    def test_generate_sf10_filename_format(self):
        """Test that generated SF10 has correct filename format"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        student = {
            'name': 'SMITH,JANE, ANN',
            'grades': {
                'LANGUAGE': 85,
                'READING AND LITERACY': 87,
                'MATH': 90,
                'GMRC': 88,
                'MAKABANSA': 86
            }
        }

        output_path = generator.generate_sf10_for_student(student, quarter=1)
        filename = os.path.basename(output_path)

        # Should contain last name and first name
        self.assertIn('SMITH', filename)
        self.assertIn('JANE', filename)
        self.assertTrue(filename.startswith('SF10_'))
        self.assertTrue(filename.endswith('.xlsx'))

    def test_generate_all_sf10s(self):
        """Test generating SF10 files for all students"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        generated_files = generator.generate_all_sf10s(quarter=1)

        # Should generate 3 files (from mock data)
        self.assertEqual(len(generated_files), 3)

        # All files should exist
        for file_path in generated_files:
            self.assertTrue(os.path.exists(file_path))

    def test_generate_sf10_different_quarters(self):
        """Test generating SF10 for different quarters"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        student = {
            'name': 'TEST,QUARTER, TWO',
            'grades': {
                'LANGUAGE': 85,
                'READING AND LITERACY': 87,
                'MATH': 90,
                'GMRC': 88,
                'MAKABANSA': 86
            }
        }

        # Generate for quarter 2
        output_path = generator.generate_sf10_for_student(student, quarter=2)
        self.assertTrue(os.path.exists(output_path))

        # Verify grade is in correct column (quarter 2 = column 11)
        wb = load_workbook(output_path)
        ws = wb.active
        # Language is row 30, quarter 2 is column 12 (1-indexed)
        language_grade = ws.cell(row=30, column=12).value
        self.assertEqual(language_grade, 85)

    def test_output_directory_creation(self):
        """Test that output directory is created if it doesn't exist"""
        non_existent_dir = os.path.join(self.test_dir, 'new_output')

        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            non_existent_dir
        )

        # Directory should be created
        self.assertTrue(os.path.exists(non_existent_dir))

    def test_student_name_parsing(self):
        """Test parsing of student names in various formats"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        test_cases = [
            {
                'name': 'LAST,FIRST, MIDDLE',
                'expected_last': 'LAST'
            },
            {
                'name': 'HYPHEN-NAME,FIRST, MIDDLE',
                'expected_last': 'HYPHEN-NAME'
            },
            {
                'name': 'LAST,FIRST SECOND, MIDDLE',
                'expected_last': 'LAST'
            }
        ]

        for test_case in test_cases:
            student = {
                'name': test_case['name'],
                'grades': {
                    'LANGUAGE': 85,
                    'READING AND LITERACY': 87,
                    'MATH': 90,
                    'GMRC': 88,
                    'MAKABANSA': 86
                }
            }

            output_path = generator.generate_sf10_for_student(student, quarter=1)
            wb = load_workbook(output_path)
            ws = wb.active

            last_name = ws.cell(row=9, column=5).value
            self.assertEqual(last_name, test_case['expected_last'])

    def test_first_name_filling(self):
        """Test that first name and middle name are filled correctly in separate fields"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        test_cases = [
            {
                'name': 'AGOT,KHIAN CLOUD, DABLO',
                'expected_last': 'AGOT',
                'expected_first': 'KHIAN CLOUD',
                'expected_middle': 'DABLO'
            },
            {
                'name': 'SMITH,JOHN, DOE',
                'expected_last': 'SMITH',
                'expected_first': 'JOHN',
                'expected_middle': 'DOE'
            },
            {
                'name': 'TEST,SINGLE, -',
                'expected_last': 'TEST',
                'expected_first': 'SINGLE',
                'expected_middle': '-'
            },
            {
                'name': 'NOmiddlename,FIRST,',
                'expected_last': 'NOmiddlename',
                'expected_first': 'FIRST',
                'expected_middle': ''
            }
        ]

        for test_case in test_cases:
            student = {
                'name': test_case['name'],
                'grades': {
                    'LANGUAGE': 85,
                    'READING AND LITERACY': 87,
                    'MATH': 90,
                    'GMRC': 88,
                    'MAKABANSA': 86
                }
            }

            output_path = generator.generate_sf10_for_student(student, quarter=1)
            wb = load_workbook(output_path)
            ws = wb.active

            # Check last name (row 9, column 5)
            last_name = ws.cell(row=9, column=5).value
            self.assertEqual(last_name, test_case['expected_last'],
                           f"Last name mismatch for {test_case['name']}")

            # Check first name (row 9, column 17)
            first_name = ws.cell(row=9, column=17).value
            self.assertEqual(first_name, test_case['expected_first'],
                           f"First name mismatch for {test_case['name']}")

            # Check middle name (row 9, column 42)
            middle_name = ws.cell(row=9, column=42).value
            # openpyxl returns None for empty cells, convert to empty string for comparison
            middle_name = middle_name if middle_name is not None else ''
            self.assertEqual(middle_name, test_case['expected_middle'],
                           f"Middle name mismatch for {test_case['name']}")

    def test_only_specified_quarter_filled(self):
        """Test that only the specified quarter is filled, others are cleared"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        student = {
            'name': 'TEST,QUARTER, ONE',
            'grades': {
                'LANGUAGE': 85,
                'READING AND LITERACY': 87,
                'MATH': 90,
                'GMRC': 88,
                'MAKABANSA': 86
            }
        }

        # Generate for quarter 1 only
        output_path = generator.generate_sf10_for_student(student, quarter=1)
        wb = load_workbook(output_path)
        ws = wb.active

        # Verify quarter 1 (column K, index 11) has grades
        self.assertEqual(ws.cell(row=30, column=11).value, 85)  # Language

        # Verify quarters 2, 3, 4 are empty (None)
        self.assertIsNone(ws.cell(row=30, column=12).value)  # Q2
        self.assertIsNone(ws.cell(row=30, column=13).value)  # Q3
        self.assertIsNone(ws.cell(row=30, column=14).value)  # Q4

    def test_other_quarters_cleared_from_template(self):
        """Test that template data in other quarters is properly cleared"""
        # Create a template with data in all quarters
        wb = Workbook()
        ws = wb.active

        # Add structure and pre-fill all quarters with dummy data
        ws.cell(row=9, column=5, value='TEMPLATE')
        ws.cell(row=30, column=2, value='Language')
        ws.cell(row=31, column=2, value='Reading and Literacy')
        ws.cell(row=32, column=2, value='Mathematics')
        ws.cell(row=33, column=2, value='GMRC (Good Manners and Right Conduct)')
        ws.cell(row=34, column=2, value='Makabansa')

        # Pre-fill all quarters with dummy data (99)
        for row in range(30, 35):
            for col in range(11, 15):  # Columns K, L, M, N
                ws.cell(row=row, column=col, value=99)

        template_with_data = os.path.join(self.test_dir, 'template_with_data.xlsx')
        wb.save(template_with_data)

        # Create generator with this template
        generator = SF10Generator(
            self.grading_sheet_path,
            template_with_data,
            self.output_dir
        )

        student = {
            'name': 'CLEAR,TEST, DATA',
            'grades': {
                'LANGUAGE': 85,
                'READING AND LITERACY': 87,
                'MATH': 90,
                'GMRC': 88,
                'MAKABANSA': 86
            }
        }

        # Generate for quarter 1
        output_path = generator.generate_sf10_for_student(student, quarter=1)
        wb_result = load_workbook(output_path)
        ws_result = wb_result.active

        # Verify Q1 has new grades
        self.assertEqual(ws_result.cell(row=30, column=11).value, 85)
        self.assertEqual(ws_result.cell(row=31, column=11).value, 87)

        # Verify Q2, Q3, Q4 are cleared (not 99 anymore)
        for row in range(30, 35):
            self.assertIsNone(ws_result.cell(row=row, column=12).value,
                            f'Q2 row {row} should be None')
            self.assertIsNone(ws_result.cell(row=row, column=13).value,
                            f'Q3 row {row} should be None')
            self.assertIsNone(ws_result.cell(row=row, column=14).value,
                            f'Q4 row {row} should be None')

    def test_single_workbook_generation(self):
        """Test generating a single workbook with multiple sheets"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        output_path = generator.generate_single_workbook_all_students(
            quarter=1,
            output_filename='Test_Single_Workbook.xlsx'
        )

        # Check file was created
        self.assertTrue(os.path.exists(output_path))

        # Load the workbook and verify sheets
        wb = load_workbook(output_path)

        # Should have 3 sheets (from mock data with 3 students)
        self.assertEqual(len(wb.sheetnames), 3)

        # Check first sheet
        first_sheet = wb[wb.sheetnames[0]]

        # Verify it has grades in Q1
        self.assertIsNotNone(first_sheet.cell(row=30, column=11).value)

        # Verify Q2/3/4 are cleared
        self.assertIsNone(first_sheet.cell(row=30, column=12).value)
        self.assertIsNone(first_sheet.cell(row=30, column=13).value)
        self.assertIsNone(first_sheet.cell(row=30, column=14).value)

    def test_single_workbook_sheet_names(self):
        """Test that sheet names are created correctly from student names"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        output_path = generator.generate_single_workbook_all_students(
            quarter=1,
            output_filename='Test_Sheet_Names.xlsx'
        )

        wb = load_workbook(output_path)

        # Check that sheet names are created
        for sheet_name in wb.sheetnames:
            # Sheet name should not be empty
            self.assertGreater(len(sheet_name), 0)

            # Sheet name should not have invalid characters
            for char in [':', '\\', '/', '?', '*', '[', ']']:
                self.assertNotIn(char, sheet_name)

            # Sheet name should be 31 characters or less (Excel limit)
            self.assertLessEqual(len(sheet_name), 31)

    def test_single_workbook_student_data(self):
        """Test that each sheet in the single workbook has correct student data"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        output_path = generator.generate_single_workbook_all_students(
            quarter=1,
            output_filename='Test_Student_Data.xlsx'
        )

        wb = load_workbook(output_path)

        # Test first sheet has data
        first_sheet = wb[wb.sheetnames[0]]

        # Check name fields are filled
        last_name = first_sheet.cell(row=9, column=5).value
        first_name = first_sheet.cell(row=9, column=17).value

        self.assertIsNotNone(last_name)
        self.assertIsNotNone(first_name)
        self.assertGreater(len(str(last_name)), 0)

        # Check grades are filled
        language_grade = first_sheet.cell(row=30, column=11).value
        self.assertIsNotNone(language_grade)
        self.assertGreater(language_grade, 0)

    def test_single_workbook_merged_cells_copied(self):
        """Test that merged cells are copied from template to each sheet"""
        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        output_path = generator.generate_single_workbook_all_students(
            quarter=1,
            output_filename='Test_Merged_Cells.xlsx'
        )

        # Load template and generated workbook
        template_wb = load_workbook(self.sf10_template_path)
        template_ws = template_wb.active
        template_merged_count = len(template_ws.merged_cells.ranges)

        generated_wb = load_workbook(output_path)

        # Check that each sheet has the same number of merged cells as template
        for sheet_name in generated_wb.sheetnames:
            ws = generated_wb[sheet_name]
            sheet_merged_count = len(ws.merged_cells.ranges)

            self.assertEqual(
                sheet_merged_count,
                template_merged_count,
                f"Sheet '{sheet_name}' has {sheet_merged_count} merged cells, "
                f"expected {template_merged_count} (same as template)"
            )

    def test_single_workbook_preserves_media_files(self):
        """Test that logos/images from template are preserved in generated workbook"""
        from zipfile import ZipFile

        generator = SF10Generator(
            self.grading_sheet_path,
            self.sf10_template_path,
            self.output_dir
        )

        output_path = generator.generate_single_workbook_all_students(
            quarter=1,
            output_filename='Test_Media_Preservation.xlsx'
        )

        # Get media files from template
        with ZipFile(self.sf10_template_path, 'r') as template_zip:
            template_media = [f for f in template_zip.namelist() if 'media/' in f]

        # Get media files from generated workbook
        with ZipFile(output_path, 'r') as generated_zip:
            generated_media = [f for f in generated_zip.namelist() if 'media/' in f]

        # Verify media files are preserved
        self.assertEqual(
            len(generated_media),
            len(template_media),
            f"Expected {len(template_media)} media files, found {len(generated_media)}"
        )

        # Verify each media file exists
        for media_file in template_media:
            self.assertIn(
                media_file,
                generated_media,
                f"Media file '{media_file}' not found in generated workbook"
            )


if __name__ == '__main__':
    unittest.main()
