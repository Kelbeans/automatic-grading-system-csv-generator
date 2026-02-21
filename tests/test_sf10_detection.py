"""
Test SF10 file detection with various filenames
"""

import unittest
import os
import tempfile
import shutil
from openpyxl import load_workbook, Workbook
from sf10_web_app import is_sf10_file
from generate_sf10 import SF10Generator


class TestSF10Detection(unittest.TestCase):
    """Test SF10 file detection by content, not filename"""

    def setUp(self):
        """Set up test fixtures"""
        self.test_dir = tempfile.mkdtemp()
        self.sf10_template = 'assets/docs/SF10.xlsx'

        # Skip tests if template not found
        if not os.path.exists(self.sf10_template):
            self.skipTest("SF10 template not found")

        # Create a mock grading sheet for testing
        self.grading_sheet = os.path.join(self.test_dir, 'test_grading.xlsx')
        self._create_mock_grading_sheet()

    def _create_mock_grading_sheet(self):
        """Create a minimal mock grading sheet for testing"""
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create SUMMARY sheet with proper name
        ws = wb.create_sheet('SUMMARY OF QUARTERLY GRADES ')

        # Add mock student data (simplified structure)
        # Row 10 is where student data starts
        students = [
            ('DOE,JOHN, A', 85, 88, 90, 87, 89),
            ('SMITH,JANE, B', 92, 90, 91, 93, 88),
            ('BROWN,BOB, C', 78, 82, 85, 80, 83),
        ]

        for idx, (name, lang, reading, math, gmrc, makabansa) in enumerate(students, start=10):
            ws.cell(row=idx, column=2, value=name)  # Column B (index 1)
            ws.cell(row=idx, column=6, value=lang)  # Column F (index 5)
            ws.cell(row=idx, column=7, value=reading)  # Column G (index 6)
            ws.cell(row=idx, column=8, value=math)  # Column H (index 7)
            ws.cell(row=idx, column=9, value=gmrc)  # Column I (index 8)
            ws.cell(row=idx, column=10, value=makabansa)  # Column J (index 9)

        wb.save(self.grading_sheet)

    def tearDown(self):
        """Clean up test fixtures"""
        if os.path.exists(self.test_dir):
            shutil.rmtree(self.test_dir)

    def test_detect_sf10_with_standard_name(self):
        """Test detection of SF10 with standard name"""
        generator = SF10Generator(
            grading_sheet_path=self.grading_sheet,
            sf10_template_path=self.sf10_template,
            output_dir=self.test_dir
        )

        # Generate a standard SF10 file
        output_path = generator.generate_single_workbook_all_students(
            quarter=1,
            output_filename='SF10_All_Students.xlsx'
        )

        # Should be detected as SF10
        self.assertTrue(is_sf10_file(output_path))

    def test_detect_sf10_with_custom_name(self):
        """Test detection of SF10 with custom/unusual name"""
        generator = SF10Generator(
            grading_sheet_path=self.grading_sheet,
            sf10_template_path=self.sf10_template,
            output_dir=self.test_dir
        )

        # Generate SF10 with various non-standard names
        test_names = [
            'MyGrades.xlsx',
            'Student_Records_2024.xlsx',
            'Q1_Complete.xlsx',
            'Grade1_DAISY.xlsx',
            'Random_Name_123.xlsx'
        ]

        for filename in test_names:
            output_path = generator.generate_single_workbook_all_students(
                quarter=1,
                output_filename=filename
            )

            # Should still be detected as SF10 regardless of name
            result = is_sf10_file(output_path)
            self.assertTrue(
                result,
                f"Failed to detect SF10 with filename: {filename}"
            )

    def test_not_detect_grading_sheet_as_sf10(self):
        """Test that grading sheets are NOT detected as SF10"""
        # The grading sheet should not be detected as SF10
        self.assertFalse(is_sf10_file(self.grading_sheet))

    def test_not_detect_template_as_sf10(self):
        """Test that the template is NOT detected as SF10 (only has 1 sheet)"""
        # The template only has 1 sheet, so should not be detected as SF10
        self.assertFalse(is_sf10_file(self.sf10_template))

    def test_detect_sf10_by_structure(self):
        """Test that detection works by checking structure, not name"""
        generator = SF10Generator(
            grading_sheet_path=self.grading_sheet,
            sf10_template_path=self.sf10_template,
            output_dir=self.test_dir
        )

        # Generate with a name that doesn't include "SF10" at all
        output_path = generator.generate_single_workbook_all_students(
            quarter=1,
            output_filename='records.xlsx'  # No "SF10" in name
        )

        # Load and verify it has the right structure
        wb = load_workbook(output_path)

        # Should have multiple sheets
        self.assertGreater(len(wb.sheetnames), 5)

        # Should have names in row 9
        first_sheet = wb[wb.sheetnames[0]]
        self.assertIsNotNone(first_sheet.cell(row=9, column=5).value)

        # Should be detected as SF10
        self.assertTrue(is_sf10_file(output_path))


if __name__ == '__main__':
    unittest.main()
